print("test")

import openpyxl

rmsign_table = str.maketrans(
    "ÁÀẢÃẠĂẮẰẲẴẶÂẤẦẨẪẬĐÈÉẺẼẸÊẾỀỂỄỆÍÌỈĨỊÓÒỎÕỌÔỐỒỔỖỘƠỚỜỞỠỢÚÙỦŨỤƯỨỪỬỮỰÝỲỶỸỴáàảãạăắằẳẵặâấầẩẫậđèéẻẽẹêếềểễệíìỉĩịóòỏõọôốồổỗộơớờởỡợúùủũụưứừửữựýỳỷỹỵ",
    "A"*17 + "D" + "E"*11 + "I"*5 + "O"*17 + "U"*11 + "Y"*5 + "a"*17 + "d" + "e"*11 + "i"*5 + "o"*17 + "u"*11 + "y"*5
)

import sys
path = sys.argv[1]
table_name = sys.argv[2]
# path = "uploads/DSDiemDanh_222SCDA331629_02CLC.xlsx"
# table_name = "SCDA331629"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row = sheet_obj.max_row 
column = sheet_obj.max_column 


#//Find Row contain table header
table = []
data_col = ['','','','','']
data_row = []
temp = 0
count = 0

for i in range(1, row + 1):  
    cell_obj = sheet_obj.cell(row = i, column = 1)  
    if(cell_obj.value == "STT" or cell_obj.value == "TT"):
        temp = i
        for j in range(1,column + 1):
            table_header = sheet_obj.cell(row = i, column = j)
            if (table_header.value != None):
                rm_sign = str(table_header.value)
                rm_sign = rm_sign.translate(rmsign_table)
                rm_sign = rm_sign.replace(" ", "_")
                table.append(rm_sign)
                print(str(table_header.value))
    if (temp!=0 and i > temp):
        n = 0
        for j in range(2,column + 1):
            cell_data = sheet_obj.cell(row = i, column = j)
            if (cell_data.value != None):
                data = str(cell_data.value)
                data = data.strip()
                data_col[n] = data
                # print(data_col)
                n+=1

        if(cell_obj.value != None):
            data_row.append(data_col)
            f = open("data.txt", "a")
            f.write("INSERT INTO " + table_name + "(" 
                + table[1] + ", " + table[2] + ", " + table[3] + ", " + table[4] + ", " + table[5]
                + ") "
                + "VALUES ("+"'" + data_row[count][0] + "'" + ", " + "'" + data_row[count][1] + "'" + ", " + "'" + data_row[count][2] + "'" + ", " + "'" + data_row[count][3] + "'" + ", " + "'" + data_row[count][4] + "'"
                + ")"+"\n")
            f.close()
            count +=1
            # print(count)
            
f = open("creatable_script.txt", "x")

f.write("CREATE TABLE " + table_name + "(" + table[0] + " MEDIUMINT NOT NULL AUTO_INCREMENT PRIMARY KEY, "
                                         + table[1] + " VARCHAR(8) NOT NULL, "
                                         + table[2] + " VARCHAR(30) NOT NULL, "
                                         + table[3] + " VARCHAR(10) NOT NULL, "
                                         + table[4] + " VARCHAR(15) NOT NULL, "
                                         + table[5] + " VARCHAR(15) NOT NULL, "
                                         + "Tuan_"  + table[6] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[7] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[8] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[9] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[10] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[11] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[12] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[13] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[14] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[15] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[16] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[17] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[18] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[19] + " VARCHAR(255) , "
                                         + "Tuan_"  + table[20] + " VARCHAR(255) "
                                         + ")")

f.close()

f = open("class_name.txt", "a")
f.write(table_name)
f.close()